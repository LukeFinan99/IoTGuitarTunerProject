import json
import os
import time
import datetime
from http.server import BaseHTTPRequestHandler, HTTPServer
import threading
import matplotlib
matplotlib.use('TkAgg')  # Use TkAgg backend for matplotlib
import matplotlib.pyplot as plt
from collections import deque
from openpyxl import Workbook, load_workbook
import csv

# Print current working directory and debug info
print(f"Current working directory: {os.getcwd()}")

# Configuration
HOST = "0.0.0.0"  # Listen on all interfaces
PORT = 8000
DATA_DIR = "guitar_tuner_data"
EXCEL_FILE = os.path.join(DATA_DIR, "tuning_data.xlsx")
CSV_FILE = os.path.join(DATA_DIR, "tuning_data.csv")
MAX_POINTS = 100

# Print paths for debugging
print(f"Data directory path: {os.path.abspath(DATA_DIR)}")
print(f"Excel file path: {os.path.abspath(EXCEL_FILE)}")
print(f"CSV file path: {os.path.abspath(CSV_FILE)}")

# Data storage
frequencies = deque(maxlen=MAX_POINTS)
statuses = deque(maxlen=MAX_POINTS)
timestamps = deque(maxlen=MAX_POINTS)

# Status mapping
STATUS_MAP = {
    "Too Low": 0,
    "In Tune": 1,
    "Too High": 2
}

# Create data directory if it doesn't exist
os.makedirs(DATA_DIR, exist_ok=True)
print(f"Created/verified data directory: {DATA_DIR}")

# Initialize CSV file
if not os.path.exists(CSV_FILE):
    with open(CSV_FILE, 'w', newline='') as file:
        writer = csv.writer(file)
        writer.writerow(["Timestamp", "Frequency (Hz)", "Tuning Status"])
    print("Initialized CSV file")

# Initialize or load Excel file
if not os.path.exists(EXCEL_FILE):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Timestamp", "Frequency (Hz)", "Tuning Status"])
    workbook.save(EXCEL_FILE)
    print("Initialized Excel file")
else:
    workbook = load_workbook(EXCEL_FILE)
    sheet = workbook.active
    print("Loaded existing Excel file")

# Real-time graph setup
plt.ion()
fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(10, 6))
print("Initialized matplotlib figures")

HTML_TEMPLATE = """<!DOCTYPE html>
<html>
<head>
    <title>Guitar Tuner Data Visualization</title>
    <script src="https://cdnjs.cloudflare.com/ajax/libs/Chart.js/3.7.0/chart.min.js"></script>
    <style>
        body { font-family: Arial, sans-serif; margin: 20px; }
        .container { max-width: 800px; margin: 0 auto; }
        .status { font-size: 24px; margin: 20px 0; text-align: center; }
        .chart-container { margin: 20px 0; }
    </style>
</head>
<body>
    <div class="container">
        <h1>Guitar Tuner Dashboard</h1>
        <div class="status">
            Current Frequency: <span id="frequency">--</span> Hz<br>
            Status: <span id="tuning-status">--</span>
        </div>
        <div class="chart-container">
            <canvas id="frequencyChart"></canvas>
        </div>
        <div class="chart-container">
            <canvas id="statusChart"></canvas>
        </div>
    </div>

    <script>
        let frequencyChart;
        let statusChart;
        
        function initCharts() {
            const freqCtx = document.getElementById('frequencyChart').getContext('2d');
            const statusCtx = document.getElementById('statusChart').getContext('2d');
            
            frequencyChart = new Chart(freqCtx, {
                type: 'line',
                data: {
                    labels: [],
                    datasets: [{
                        label: 'Frequency (Hz)',
                        data: [],
                        borderColor: 'rgb(75, 192, 192)',
                        tension: 0.1
                    }]
                },
                options: {
                    responsive: true,
                    scales: {
                        y: {
                            beginAtZero: false
                        }
                    }
                }
            });
            
            statusChart = new Chart(statusCtx, {
                type: 'line',
                data: {
                    labels: [],
                    datasets: [{
                        label: 'Tuning Status',
                        data: [],
                        borderColor: 'rgb(153, 102, 255)',
                        tension: 0.1
                    }]
                },
                options: {
                    responsive: true,
                    scales: {
                        y: {
                            min: -1,
                            max: 3,
                            ticks: {
                                callback: function(value) {
                                    return ['Too Low', 'In Tune', 'Too High'][value];
                                }
                            }
                        }
                    }
                }
            });
        }

        function updateCharts() {
            fetch('/data')
                .then(response => response.json())
                .then(data => {
                    if (data.current) {
                        document.getElementById('frequency').textContent = 
                            data.current.frequency.toFixed(1);
                        document.getElementById('tuning-status').textContent = 
                            data.current.status;
                    }
                    
                    const timestamps = data.history.map(d => d.timestamp);
                    const frequencies = data.history.map(d => d.frequency);
                    const statuses = data.history.map(d => d.status_value);
                    
                    frequencyChart.data.labels = timestamps;
                    frequencyChart.data.datasets[0].data = frequencies;
                    frequencyChart.update();
                    
                    statusChart.data.labels = timestamps;
                    statusChart.data.datasets[0].data = statuses;
                    statusChart.update();
                });
        }

        // Initialize charts when page loads
        initCharts();
        
        // Update every second
        setInterval(updateCharts, 1000);
    </script>
</body>
</html>
"""

def update_plot():
    """Update the graph with current data."""
    try:
        ax1.clear()
        ax2.clear()

        # Plot frequency over time
        ax1.plot(range(len(frequencies)), list(frequencies), '-o', label="Frequency (Hz)")
        ax1.set_title("Frequency Over Time")
        ax1.set_ylabel("Frequency (Hz)")
        ax1.legend()
        ax1.grid()

        # Convert status strings to numerical values for plotting
        status_values = [STATUS_MAP.get(status, -1) for status in statuses]
        
        # Plot status over time
        ax2.plot(range(len(status_values)), status_values, '-o', label="Tuning Status")
        ax2.set_title("Tuning Status")
        ax2.set_ylabel("Status (0=Too Low, 1=In Tune, 2=Too High)")
        ax2.legend()
        ax2.grid()

        plt.tight_layout()
        plt.draw()
        plt.pause(0.1)
    except Exception as e:
        print(f"Error updating plot: {str(e)}")

class RequestHandler(BaseHTTPRequestHandler):
    def do_GET(self):
        print(f"Received GET request for path: {self.path}")
        if self.path == '/':
            self.send_response(200)
            self.send_header('Content-Type', 'text/html')
            self.end_headers()
            self.wfile.write(HTML_TEMPLATE.encode())
            print("Served HTML template")
        
        elif self.path == '/data':
            self.send_response(200)
            self.send_header('Content-Type', 'application/json')
            self.send_header('Access-Control-Allow-Origin', '*')
            self.end_headers()
            
            current = {}
            if timestamps:
                status_str = statuses[-1]
                status_value = STATUS_MAP.get(status_str, -1)
                current = {
                    'timestamp': timestamps[-1].strftime('%Y-%m-%d %H:%M:%S'),
                    'frequency': frequencies[-1],
                    'status': status_str,
                    'status_value': status_value
                }
            
            history = []
            for t, f, s in zip(reversed(timestamps), reversed(frequencies), reversed(statuses)):
                status_value = STATUS_MAP.get(s, -1)
                history.append({
                    'timestamp': t.strftime('%Y-%m-%d %H:%M:%S'),
                    'frequency': f,
                    'status': s,
                    'status_value': status_value
                })
            
            response_data = {
                'current': current,
                'history': history
            }
            
            self.wfile.write(json.dumps(response_data).encode())
            print("Served data JSON")
    
    def do_POST(self):
        print("\n--- RECEIVED POST REQUEST ---")
        
        # Get content length and read data
        content_length = int(self.headers.get("Content-Length", 0))
        print(f"Content Length: {content_length}")
        
        post_data = self.rfile.read(content_length).decode("utf-8")
        print(f"Raw POST Data: {post_data}")
        
        try:
            # Parse JSON data with more robust error handling
            data = json.loads(post_data)
            print(f"Parsed JSON Data: {data}")
            
            frequency = float(data.get("frequency", 0))
            status = data.get("status", "Unknown")
            timestamp = datetime.datetime.now()

            # Detailed logging of processed data
            print(f"Processing Data:")
            print(f"  Timestamp: {timestamp}")
            print(f"  Frequency: {frequency} Hz")
            print(f"  Status: {status}")
            
            # Store data
            frequencies.append(frequency)
            statuses.append(status)
            timestamps.append(timestamp)

            # Save to Excel and CSV with error handling
            try:
                sheet.append([timestamp, frequency, status])
                workbook.save(EXCEL_FILE)
                print("Data saved to Excel successfully")

                with open(CSV_FILE, 'a', newline='') as file:
                    writer = csv.writer(file)
                    writer.writerow([timestamp, frequency, status])
                print("Data saved to CSV successfully")
            except Exception as save_error:
                print(f"Error saving data: {save_error}")

            # Respond with success
            self.send_response(200)
            self.send_header('Content-Type', 'application/json')
            self.send_header('Access-Control-Allow-Origin', '*')
            self.end_headers()
            
            response = json.dumps({"status": "Data received successfully"})
            self.wfile.write(response.encode())
            
            print("--- POST REQUEST PROCESSED SUCCESSFULLY ---\n")

        except json.JSONDecodeError:
            print("Error: Invalid JSON data")
            self.send_response(400)
            self.end_headers()
            self.wfile.write(b"Invalid JSON")
        except Exception as e:
            print(f"Unexpected error processing POST request: {e}")
            import traceback
            traceback.print_exc()
            self.send_response(500)
            self.end_headers()
            self.wfile.write(b"Server error")

def start_server():
    server = HTTPServer((HOST, PORT), RequestHandler)
    print(f"Server started at http://{HOST}:{PORT}")
    server.serve_forever()

if __name__ == "__main__":
    print("Starting guitar tuner server...")
    
    # Start server in a separate thread
    server_thread = threading.Thread(target=start_server, daemon=True)
    server_thread.start()

    print("Server running. Waiting for data...")
    try:
        while True:
            if len(frequencies) > 0:
                update_plot()
            time.sleep(0.1)
    except KeyboardInterrupt:
        print("\nShutting down server...")